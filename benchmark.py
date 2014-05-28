# -*- coding: utf-8 -*-
import argparse
from functools import wraps
import re
import timeit
from itertools import cycle, izip

ROWS = 1000
COLUMNS = 100
RUN_COUNT = 10

VALUES = cycle([1, None, "foobar", 2.32])


def skip(description):
    def skip_decorator(fun):

        @wraps(fun)
        def wrapped(*args, **kwargs):
            fun(*args, **kwargs)

        wrapped.skip = description

        return wrapped

    return skip_decorator


def get_benchmarks():
    return [
        item
        for item in globals().values()
        if callable(item) and item.__name__.startswith('benchmark')
    ]


def benchmark_xlwt():
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('A Test Sheet')

    for row, value in izip(xrange(ROWS), VALUES):
        for column in xrange(COLUMNS):
            sheet.write(row, column, value)

    workbook.save('benchmark_xlwt.xlsx')


def benchmark_xlsxcessive():
    from xlsxcessive import xlsx as xcessive

    workbook = xcessive.Workbook()
    sheet = workbook.new_sheet('Sheet1')

    for row, value in izip(xrange(ROWS), VALUES):
        for column in xrange(COLUMNS):
            sheet.cell(coords=(row, column), value=value)

    xcessive.save(workbook, 'benchmark_xlsxcessive.xlsx')


@skip('ooxml is FUBAR')
def benchmark_ooxml():
    from ooxml import spreadsheet as ooxmls

    workbook = ooxmls.Spreadsheet()
    sheet = workbook.sheet(1)

    for row, value in izip(xrange(ROWS), VALUES):
        for column in xrange(COLUMNS):
            sheet.set_cell(column, row, value)

    workbook.save('benchmark_ooxml.xlsx')


def benchmark_openpyxl_rows():
    """OpenPyXL using sheet.append('row')

    As documentation says this should be faster:
    http://pythonhosted.org/openpyxl/optimized.html#optimized-writer
    """

    import openpyxl

    workbook = openpyxl.workbook.Workbook(optimized_write=True)
    sheet = workbook.create_sheet()
    sheet.title = 'Sheet1'
    # note: pyopenxl indexes rows and columns starting from 1
    for row, value in izip(xrange(1, ROWS + 1), VALUES):
        sheet.append([str(value) for _ in xrange(1, COLUMNS + 1)])

    workbook.save('benchmark_openpyxl_rows.xslx')


def benchmark_openpyxl():
    """OpenPyXL using sheet.cell().value = value"""
    import openpyxl
    from openpyxl.cell import get_column_letter

    workbook = openpyxl.workbook.Workbook(optimized_write=True)
    sheet = workbook.create_sheet()
    sheet.title = 'Sheet1'

    # note: pyopenxl indexes rows and columns starting from 1
    for row, value in izip(xrange(1, ROWS + 1), VALUES):
        for column in xrange(1, COLUMNS + 1):
            sheet.cell('%s%s' % (get_column_letter(column), row)).value = value

    workbook.save('benchmark_openpyxl.xslx')


def benchmark_pyexcelerate():
    import pyexcelerate

    workbook = pyexcelerate.Workbook()

    data = [
        [value for column in xrange(COLUMNS)]
         for __, value in izip(xrange(ROWS), VALUES)
    ]

    workbook.new_sheet('Test 1', data=data)
    workbook.save('benchmark_pyexcelerate.xlsx')


def benchmark_excellent():
    from collections import OrderedDict
    from excellent import Writer, XL

    output = open("benchmark_excellent.xls", "wb")
    sheet_manager = XL()

    excel = Writer(sheet_manager, output)

    sheet_manager.use_sheet("Sheet1")

    excel.write((
        OrderedDict((
            (str(column), value)
            for column in xrange(COLUMNS)
        ))
        for row, value
        in izip(xrange(ROWS), VALUES)
    ))

    excel.save()


def benchmark_xlsxwriter():
    import xlsxwriter
    workbook = xlsxwriter.Workbook('benchmark_xlsxwriter.xlsx')
    sheet = workbook.add_worksheet()

    for row, value in izip(xrange(ROWS), VALUES):
        for column in xrange(COLUMNS):
            sheet.write(column, row, value)

    workbook.close()


def benchmark_csv():
    import csv

    with open('benchmark_scv.csv', 'wb') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows([
            [value for _ in xrange(COLUMNS)]
            for row, value in izip(xrange(ROWS), VALUES)
        ])


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Simple benchmark for various Excel/XSLX python libraries'
    )
    parser.add_argument(
        '--filter', '-f',
        metavar='regex', type=re.compile,
        default=re.compile(''),
        help='benchmark filter',
    )

    parser.add_argument(
        '--tests', '-t',
        metavar='number', type=int,
        default=10,
        help='number of test runs',
    )

    parser.add_argument(
        '--columns', '-c',
        metavar='number', type=int,
        default=100,
        help='number of test spreadsheed columns',
    )

    parser.add_argument(
        '--rows', '-r',
        metavar='number', type=int,
        default=1000,
        help='number of test spreadsheet rows',
    )

    args = parser.parse_args()

    for stmt in sorted(filter(lambda fun: args.filter.search(fun.__name__),
                              get_benchmarks()),
                       key=lambda fun: fun.__name__):
        if hasattr(stmt, 'skip'):
            print("# SKIP {0} ({1})".format(stmt.__name__, stmt.skip))
            continue

        # do it globally to simplify test running and avoid
        # passing parameters to benchmark_function
        # (timeit - I hate you!)
        COLUMNS = args.columns
        ROWS = args.rows

        timer = timeit.Timer(stmt, 'gc.enable()')
        try:
            result = timer.timeit(number=args.tests)
        except ImportError, err:
            print("# SKIP {0} ({1})".format(stmt.__name__, err))
            continue

        print("{0:30} {1:5f}".format(stmt.__name__, result/args.tests,))
